import os
import io
from datetime import datetime, date
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Sum, Avg, Count, Q
from projects.models import Project
from sales.models import SaleBooking, Payment
from material_management.models import Material, MaterialRequisition
from accounts.models import JournalEntry, Budget
import json

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def get_report_data(report_type, user, date_from=None, date_to=None, project_ids=None):
    """Get data for different report types"""
    data = {
        'report_type': report_type,
        'generated_at': timezone.now(),
        'date_from': date_from,
        'date_to': date_to,
    }
    
    # Base queryset for user's projects
    projects_qs = Project.objects.filter(created_by=user)
    if project_ids:
        projects_qs = projects_qs.filter(id__in=project_ids)
    if date_from:
        projects_qs = projects_qs.filter(created_at__gte=date_from)
    if date_to:
        projects_qs = projects_qs.filter(created_at__lte=date_to)
    
    projects = list(projects_qs)
    
    if report_type == 'project_summary':
        data.update({
            'total_projects': projects_qs.count(),
            'active_projects': projects_qs.filter(status='in_progress').count(),
            'completed_projects': projects_qs.filter(status='completed').count(),
            'total_budget': projects_qs.aggregate(Sum('budget'))['budget__sum'] or 0,
            'total_spent': projects_qs.aggregate(Sum('actual_cost'))['actual_cost__sum'] or 0,
            'projects': projects,
        })
    
    elif report_type == 'financial_overview':
        total_budget = projects_qs.aggregate(Sum('budget'))['budget__sum'] or 0
        total_spent = projects_qs.aggregate(Sum('actual_cost'))['actual_cost__sum'] or 0
        data.update({
            'total_budget': total_budget,
            'total_spent': total_spent,
            'remaining_budget': total_budget - total_spent,
            'budget_utilization': (total_spent / total_budget * 100) if total_budget > 0 else 0,
            'projects': projects,
        })
    
    elif report_type == 'project_progress':
        data.update({
            'projects': projects,
            'avg_progress': projects_qs.aggregate(Avg('progress'))['progress__avg'] or 0,
            'on_time_projects': projects_qs.filter(progress__gte=80).count(),
            'behind_schedule': projects_qs.filter(progress__lt=50).count(),
        })
    
    elif report_type == 'material_usage':
        materials = Material.objects.filter(project__in=projects_qs)
        data.update({
            'materials': materials,
            'total_materials': materials.count(),
            'total_cost': materials.aggregate(Sum('cost'))['cost__sum'] or 0,
            'projects': projects,
        })
    
    elif report_type == 'cost_analysis':
        data.update({
            'projects': projects,
            'budget_variance': sum((p.actual_cost - p.budget) for p in projects if p.budget and p.actual_cost),
            'cost_per_project': sum(p.actual_cost for p in projects) / len(projects) if projects else 0,
        })
    
    return data

def generate_pdf_report(report_data, report_request):
    """Generate PDF report using ReportLab"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("ReportLab is not installed. Install with: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#7c3aed')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor('#7c3aed')
    )
    
    # Build content
    story = []
    
    # Title
    story.append(Paragraph(f"{report_request.get_report_type_display()}", title_style))
    story.append(Paragraph(f"Generated on: {report_data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Report content based on type
    if report_data['report_type'] == 'project_summary':
        story.append(Paragraph("Project Summary", heading_style))
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Projects', str(report_data['total_projects'])],
            ['Active Projects', str(report_data['active_projects'])],
            ['Completed Projects', str(report_data['completed_projects'])],
            ['Total Budget', f"${report_data['total_budget']:,.2f}"],
            ['Total Spent', f"${report_data['total_spent']:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Projects list
        if report_data['projects']:
            story.append(Paragraph("Projects Details", heading_style))
            project_data = [['Project Name', 'Status', 'Progress', 'Budget', 'Actual Cost']]
            
            for project in report_data['projects']:
                project_data.append([
                    project.name,
                    project.get_status_display(),
                    f"{project.progress}%",
                    f"${project.budget:,.2f}" if project.budget else "N/A",
                    f"${project.actual_cost:,.2f}" if project.actual_cost else "N/A"
                ])
            
            project_table = Table(project_data, colWidths=[1.5*inch, 1*inch, 0.8*inch, 1.2*inch, 1.2*inch])
            project_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            story.append(project_table)
    
    elif report_data['report_type'] == 'financial_overview':
        story.append(Paragraph("Financial Overview", heading_style))
        
        financial_data = [
            ['Financial Metric', 'Amount'],
            ['Total Budget', f"${report_data['total_budget']:,.2f}"],
            ['Total Spent', f"${report_data['total_spent']:,.2f}"],
            ['Remaining Budget', f"${report_data['remaining_budget']:,.2f}"],
            ['Budget Utilization', f"{report_data['budget_utilization']:.1f}%"],
        ]
        
        financial_table = Table(financial_data, colWidths=[2*inch, 2*inch])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(financial_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(report_data, report_request):
    """Generate Excel report using openpyxl"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("OpenPyXL is not installed. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = f"{report_request.get_report_type_display()}"
    ws['A1'].font = Font(bold=True, size=16, color="7c3aed")
    ws['A2'] = f"Generated on: {report_data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(size=10)
    
    current_row = 4
    
    if report_data['report_type'] == 'project_summary':
        # Summary section
        ws[f'A{current_row}'] = "Project Summary"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="7c3aed")
        current_row += 2
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Projects', report_data['total_projects']],
            ['Active Projects', report_data['active_projects']],
            ['Completed Projects', report_data['completed_projects']],
            ['Total Budget', report_data['total_budget']],
            ['Total Spent', report_data['total_spent']],
        ]
        
        for row_data in summary_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if current_row == 4:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
            current_row += 1
        
        current_row += 2
        
        # Projects details
        if report_data['projects']:
            ws[f'A{current_row}'] = "Projects Details"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="7c3aed")
            current_row += 2
            
            # Headers
            headers = ['Project Name', 'Status', 'Progress (%)', 'Budget', 'Actual Cost']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            current_row += 1
            
            # Data rows
            for project in report_data['projects']:
                row_data = [
                    project.name,
                    project.get_status_display(),
                    project.progress,
                    project.budget or 0,
                    project.actual_cost or 0
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1
    
    elif report_data['report_type'] == 'financial_overview':
        ws[f'A{current_row}'] = "Financial Overview"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="7c3aed")
        current_row += 2
        
        financial_data = [
            ['Financial Metric', 'Amount'],
            ['Total Budget', report_data['total_budget']],
            ['Total Spent', report_data['total_spent']],
            ['Remaining Budget', report_data['remaining_budget']],
            ['Budget Utilization (%)', report_data['budget_utilization']],
        ]
        
        for row_data in financial_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if current_row == 4:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
            current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_csv_report(report_data, report_request):
    """Generate CSV report"""
    import csv
    
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    
    # Header
    writer.writerow([f"{report_request.get_report_type_display()}"])
    writer.writerow([f"Generated on: {report_data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}"])
    writer.writerow([])
    
    if report_data['report_type'] == 'project_summary':
        writer.writerow(['Project Summary'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Projects', report_data['total_projects']])
        writer.writerow(['Active Projects', report_data['active_projects']])
        writer.writerow(['Completed Projects', report_data['completed_projects']])
        writer.writerow(['Total Budget', report_data['total_budget']])
        writer.writerow(['Total Spent', report_data['total_spent']])
        writer.writerow([])
        
        if report_data['projects']:
            writer.writerow(['Projects Details'])
            writer.writerow(['Project Name', 'Status', 'Progress (%)', 'Budget', 'Actual Cost'])
            for project in report_data['projects']:
                writer.writerow([
                    project.name,
                    project.get_status_display(),
                    project.progress,
                    project.budget or 0,
                    project.actual_cost or 0
                ])
    
    elif report_data['report_type'] == 'financial_overview':
        writer.writerow(['Financial Overview'])
        writer.writerow(['Financial Metric', 'Amount'])
        writer.writerow(['Total Budget', report_data['total_budget']])
        writer.writerow(['Total Spent', report_data['total_spent']])
        writer.writerow(['Remaining Budget', report_data['remaining_budget']])
        writer.writerow(['Budget Utilization (%)', report_data['budget_utilization']])
    
    buffer.seek(0)
    return buffer
